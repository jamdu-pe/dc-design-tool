"""시나리오 비교 산출물·CLI 테스트."""
import pathlib

import yaml
from openpyxl import load_workbook
from typer.testing import CliRunner

from dc_design_tool.cli import app
from dc_design_tool.engine import scenario
from dc_design_tool.reports.comparison_xlsx import write_comparison

runner = CliRunner()

SWEEP = {
    "project": "GB200 vs GB300 비교",
    "base": {"it_power_mw": 5.0, "rack_id": "nvidia_gb200_nvl72", "tier": "III",
             "electrical_redundancy": "N+1", "mechanical_redundancy": "N+1",
             "target_pue": 1.6},
    "sweep": {"rack_id": ["nvidia_gb200_nvl72", "nvidia_gb300_nvl72"],
              "electrical_redundancy": ["N+1", "2N"]},
}


def _file(tmp_path: pathlib.Path, data: dict) -> str:
    p = tmp_path / "sweep.yaml"
    p.write_text(yaml.safe_dump(data, allow_unicode=True), encoding="utf-8")
    return str(p)


# ---------- 산출물 ----------

def test_comparison_xlsx_has_one_row_per_scenario(tmp_path):
    rows = scenario.run_sweep(SWEEP["base"], SWEEP["sweep"])
    wb = load_workbook(write_comparison(rows, "비교", str(tmp_path)))
    assert "시나리오비교" in wb.sheetnames
    values = list(wb["시나리오비교"].values)
    assert len(values) == len(rows) + 1          # 헤더 + 시나리오
    assert values[0][0] == "시나리오"


def test_comparison_xlsx_includes_capex_absence_note(tmp_path):
    rows = scenario.run_sweep(SWEEP["base"], {})
    path = write_comparison(rows, "비교", str(tmp_path))
    wb = load_workbook(path)
    text = "\n".join(str(c) for r in wb["시나리오비교"].values for c in r if c)
    assert "비용" in text


# ---------- CLI ----------

def test_compare_command_writes_xlsx_and_prints_table(tmp_path):
    out = tmp_path / "out"
    res = runner.invoke(app, ["compare", "--spec", _file(tmp_path, SWEEP),
                              "--out", str(out)])
    assert res.exit_code == 0, res.output
    assert (out / "시나리오비교.xlsx").is_file()
    assert "nvidia_gb300_nvl72" in res.output
    assert "시나리오" in res.output


def test_compare_sorts_by_requested_metric(tmp_path):
    res = runner.invoke(app, ["compare", "--spec", _file(tmp_path, SWEEP),
                              "--out", str(tmp_path / "out"),
                              "--sort", "total_building_m2"])
    assert res.exit_code == 0, res.output


def test_compare_rejects_unknown_sort_metric(tmp_path):
    res = runner.invoke(app, ["compare", "--spec", _file(tmp_path, SWEEP),
                              "--out", str(tmp_path / "out"), "--sort", "bogus"])
    assert res.exit_code != 0
    assert "비교 지표" in res.output


def test_compare_rejects_unknown_sweep_key(tmp_path):
    bad = {**SWEEP, "sweep": {"nope": [1, 2]}}
    res = runner.invoke(app, ["compare", "--spec", _file(tmp_path, bad),
                              "--out", str(tmp_path / "out")])
    assert res.exit_code != 0
    assert "스윕 키" in res.output


def test_compare_reports_failed_scenario_without_aborting(tmp_path):
    data = {**SWEEP, "sweep": {"rack_id": ["nvidia_gb200_nvl72", "ghost_rack"]}}
    res = runner.invoke(app, ["compare", "--spec", _file(tmp_path, data),
                              "--out", str(tmp_path / "out")])
    assert res.exit_code == 0, res.output
    assert "카탈로그 부재" in res.output


def test_example_sweep_file_runs(tmp_path):
    spec = pathlib.Path("examples/sweep_chip_generation.yaml")
    assert spec.is_file(), "예제 스윕 파일 없음"
    res = runner.invoke(app, ["compare", "--spec", str(spec), "--out", str(tmp_path)])
    assert res.exit_code == 0, res.output
