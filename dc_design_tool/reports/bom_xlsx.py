"""BOM/부하 요약 Excel 산출."""
from __future__ import annotations
import pathlib
from openpyxl import Workbook
from openpyxl.styles import Font
from ..engine.models import SizingResult


def write_bom(result: SizingResult, out_dir: str) -> str:
    out = pathlib.Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    s = wb.active
    s.title = "요약"
    s["A1"] = f"프로젝트: {result.project}"
    s["A1"].font = Font(bold=True, size=13)
    rows = [
        ("랙 모델", result.rack_id), ("랙 수량", result.rack_count),
        ("총 IT부하(kW)", result.it_power_kw),
        ("액냉 열량(kW)", result.cooling["liquid_kw"]),
        ("냉각수 유량(L/min)", result.cooling["coolant_flow_lpm"]),
        ("총 냉동톤(RT)", result.cooling["total_rt"]),
        ("UPS 수량", result.electrical["ups_qty"]),
        ("발전기 수량", result.electrical["generator_qty"]),
        ("PUE(추정)", result.electrical["pue_estimate"]),
        ("화이트스페이스(m2)", result.space["white_space_m2"]),
        ("총 건축면적(m2)", result.space["total_building_m2"]),
        ("규격 위반 건수", len(result.compliance.violations) if result.compliance else 0),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        s[f"A{i}"], s[f"B{i}"] = k, v

    b = wb.create_sheet("BOM")
    hdr = ["도메인", "품목", "모델", "단위용량", "수량", "비고"]
    b.append(hdr)
    for c in b[1]:
        c.font = Font(bold=True)
    for li in result.bom:
        b.append([li.domain, li.item, li.model, li.unit_capacity, li.qty, li.note])

    if result.compliance:
        cs = wb.create_sheet("규격검증")
        cs.append(["심각도", "코드", "도메인", "판정", "설계값", "요구값", "근거"])
        for c in cs[1]:
            c.font = Font(bold=True)
        order = {"violation": 0, "warning": 1, "info": 2}
        label = {"violation": "위반", "warning": "경고", "info": "정보"}
        for f in sorted(result.compliance.findings, key=lambda x: order[x.severity]):
            cs.append([label[f.severity], f.code, f.domain, f.message,
                       f.actual, f.required, f.rule])

    n = wb.create_sheet("가정_경고")
    n.append(["구분", "내용"])
    for a in result.assumptions:
        n.append(["가정", a])
    for w in result.warnings:
        n.append(["경고", w])

    path = out / "BOM_부하요약.xlsx"
    wb.save(path)
    return str(path)
